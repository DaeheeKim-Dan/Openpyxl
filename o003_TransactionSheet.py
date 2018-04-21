from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import openpyxl as ox
from openpyxl import Workbook


def style_cell(ws, cell_range, border=Border(), alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    rows = ws[cell_range]

    for row in rows:
        for cell in row:
            cell.border = cell.border + top + right
    for cell in rows[-1]:
        cell.border = cell.border + bottom + right

        
def style_range(ws, cell_range, border=Border(), alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    rows = ws[cell_range]

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right


fileName='o003_TransactionSheet.xlsx'

wb=ox.load_workbook(fileName)
sht=wb.active


# Code for Merged_Cell_Range
MC=sht.merged_cells.ranges                  # Merged Cell CellRange List

# Code for getting Merged Cell Ranges list
# The result will looks like below.
# ['B48:F49', 'G48:L49', ..., 'B1:E2']
# MCR=str(sht.merged_cells).split(' ')
MCR = [item.coord for item in MC]           # Merged Cell Range String List


# Code for extracting the first cell address of each merged cell
# The result will looks like below as per above result.
# ['B48', 'G48', ..., 'B1']
MCFC= [item.split(':')[0] for item in MCR]  # Merged Cell Range First Cell
MCFB= [item.bounds[-1] for item in MC]      # Merged Cell Range Bound (Starting Row)


# Code for printing the content of each merged cell
# The result will be like below
# 인 수 자
# =IF(G23="","",G23)
# ...
# 거래일자
# for j in range(len(MC)):
#     print(sht[MCFC[j]].value)

# Code for get defined name list
# 
nameList=[item.name for item in wb.defined_names.definedName]

# Set date information for the defined name range 'DATE'
adr=tuple(wb.defined_names['DATE'].destinations)
sht[adr[-1][-1]].value='2018-04-20'
# Todo : try catch expression is needed.

thick_blue=['B2:AI3', 'B4:R11', 'S4:AI11', 'B12:AI22', 'B23:AI24']
thick_brown=['B27:AI28', 'B29:R36', 'S29:AI36', 'B37:AI47', 'B48:AI49']

customer_thick= Side(border_style="thick", color="0000ff")
supplier_thick = Side(border_style="thick", color="953737")
# font = Font(b=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")

for item_blue in thick_blue:
    border_blue_thk=Border(top=customer_thick, left=customer_thick, right=customer_thick, bottom=customer_thick)
    style_range(sht, item_blue, border=border_blue_thk, alignment=al)

for item_brown in thick_brown:
    border_brown_thk=Border(top=supplier_thick, left=supplier_thick, right=supplier_thick, bottom=supplier_thick)
    style_range(sht, item_brown, border=border_brown_thk, alignment=al)

    
customer = Side(border_style="thin", color="0000ff") # Blue Border
supplier = Side(border_style="thin", color="953737") # Brown Border


border = Border(top=customer, left=customer, right=customer, bottom=customer)
style_cell(sht, 'B12:C22', border=border, alignment=al)
border = Border(top=supplier, left=supplier, right=supplier, bottom=supplier)
style_cell(sht, 'B37:C47', border=border, alignment=al)


for i, item in enumerate(MC):
    if item.bounds[1]>25:
        border = Border(top=supplier, left=supplier, right=supplier, bottom=supplier)
        style_range(sht, MCR[i], border=border, alignment=al)
    else:
        border = Border(top=customer, left=customer, right=customer, bottom=customer)
        style_range(sht, MCR[i], border=border, alignment=al)

    
wb.save(fileName)
